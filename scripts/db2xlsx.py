#!/usr/bin/env python3

import sqlite3
import pandas as pd
import argparse
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

CHUNK_SIZE = 10000  # You can tune this based on memory limits

def export_table_chunked(conn, table_name, writer_path):
    print(f"Exporting table (in chunks): {table_name}")

    offset = 0
    is_first_chunk = True

    while True:
        query = f"SELECT * FROM {table_name} LIMIT {CHUNK_SIZE} OFFSET {offset}"
        df_chunk = pd.read_sql_query(query, conn)

        if df_chunk.empty:
            break

        if is_first_chunk:
            wb = Workbook()
            ws = wb.active
            ws.title = table_name
            is_first_chunk = False
        else:
            wb = load_workbook(writer_path)
            if table_name in wb.sheetnames:
                ws = wb[table_name]
            else:
                ws = wb.create_sheet(title=table_name)

        # Write header only on the first chunk
        include_header = offset == 0
        for row in dataframe_to_rows(df_chunk, index=False, header=include_header):
            ws.append(row)

        wb.save(writer_path)
        wb.close()

        offset += CHUNK_SIZE

def convert_db_to_xlsx_chunked(db_path, xlsx_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()

    if not tables:
        print("No tables found in the database.")
        return

    # Delete existing Excel file if it exists
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)

    for table_name_tuple in tables:
        table_name = table_name_tuple[0]
        export_table_chunked(conn, table_name, xlsx_path)

    print(f"\n✅ Conversion complete! Excel file saved to: {xlsx_path}")
    conn.close()

def main():
    parser = argparse.ArgumentParser(description="Convert a SQLite .db file to Excel (.xlsx) in chunks")
    parser.add_argument("db_file", help="Path to the SQLite .db file")
    parser.add_argument("-o", "--output", help="Output Excel file (.xlsx)", default=None)

    args = parser.parse_args()
    db_path = args.db_file
    xlsx_path = args.output if args.output else os.path.splitext(db_path)[0] + ".xlsx"

    if not os.path.isfile(db_path):
        print(f"❌ Error: File '{db_path}' does not exist.")
        return

    convert_db_to_xlsx_chunked(db_path, xlsx_path)

if __name__ == "__main__":
    main()
